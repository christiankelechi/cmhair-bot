import io
import csv
import logging
import uuid
import openpyxl
from telegram import Update
from telegram.ext import (
    CommandHandler, ContextTypes, ConversationHandler, MessageHandler, filters
)
from handlers.base import BaseHandler
import api

log = logging.getLogger(__name__)

WAIT_EXCEL = 1

class ExcelProductHandler(BaseHandler):
    
    async def start(self, update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
        if not await self.require_auth(update, ctx):
            return ConversationHandler.END
            
        ctx.user_data['pending_images'] = []
            
        await update.message.reply_text(
            "🛍️ *Bulk Add Products via Excel*\n\n"
            "This wizard lets you upload multiple products safely.\n\n"
            "**Need to add images?**\n"
            "Simple! Just send your product images directly to me in this chat. Then, upload your filled out Excel sheet. I will automatically attach all the images you sent to the products in your sheet! 🪄\n\n"
            "Send /cancel to abort.",
            parse_mode="Markdown"
        )
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products Template"
        
        headers = [
            "Product Name", "Slug", "Product Code", "Original Price", 
            "Discount Price", "Stock", "Capsize", "Inches", "Unavailable Lengths", 
            "Bundles", "Color", "Parting", "Styling", "Category", 
            "Image Color Mapping", "Description"
        ]
        ws.append(headers)
        
        # Provide dummy data correctly aligned with the project schema
        dummy_row = [
            "Bone Straight Closure Wig", "bone-straight-closure-wig", "BSC-01", 150.0,
            120.0, 50, "Medium", "14:$120, 16:$150, 18:$170", "10, 12",
            "3", "Natural Black", "Middle Part", "Straight", "Wigs",
            '"Natural Black":"1"',
            "Premium quality 100% human hair bone straight closure wig."
        ]
        ws.append(dummy_row)
        
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        out.name = "CMHair_Product_Template.xlsx"
        
        await update.message.reply_document(
            document=out,
            caption="📂 *Here is your Product Template!*\n\n"
                    "1. Send me any images you want to use for these products.\n"
                    "2. Upload your filled-out Excel file (you can modify or delete the sample row).\n"
                    "3. I'll automatically join them together! (Send /cancel at any time)",
            parse_mode="Markdown"
        )
        
        return WAIT_EXCEL

    def extract_value(self, row, headers, possible_keys):
        """Helper to find a value in a row given multiple possible header names."""
        for key in possible_keys:
            for idx, header in enumerate(headers):
                if header and key.lower() in str(header).lower():
                    if idx < len(row):
                        val = row[idx]
                        if val is not None and str(val).strip() != "":
                            return str(val).strip()
        return None
        
    def split_commas_or_semicolons(self, val):
        if not val:
            return []
        if ';' in val:
            return [x.strip() for x in val.split(';') if x.strip()]
        return [x.strip() for x in val.split(',') if x.strip()]

    async def handle_photo(self, update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
        try:
            if 'pending_images' not in ctx.user_data:
                ctx.user_data['pending_images'] = []
                
            photo_file = await update.message.photo[-1].get_file()
            file_content = await photo_file.download_as_bytearray()
            ctx.user_data['pending_images'].append(file_content)
            
            count = len(ctx.user_data['pending_images'])
            await update.message.reply_text(f"✅ Queued Image #{count}. (Send more, or upload your layout file!)")
        except Exception as e:
            log.error("Failed to receive image bytes: %s", e)
            
        return WAIT_EXCEL

    async def handle_document(self, update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
        doc = update.message.document
        if not (doc.file_name.endswith('.xlsx') or doc.file_name.endswith('.csv')):
            await update.message.reply_text("❌ Please send a valid `.xlsx` or `.csv` file, or a Photo to get an image link. Or /cancel.")
            return WAIT_EXCEL
            
        msg = await update.message.reply_text("⏳ Processing file and images...")
        try:
            uploaded_image_urls = []
            pending_images = ctx.user_data.get('pending_images', [])
            
            if pending_images:
                await msg.edit_text(f"⏳ Uploading {len(pending_images)} attached image(s)...")
                for img_bytes in pending_images:
                    try:
                        filename = f"image_{uuid.uuid4().hex[:8]}.jpg"
                        url = await api.upload_image(img_bytes, self.token(ctx), filename)
                        uploaded_image_urls.append(url)
                    except Exception as e:
                        log.error("Failed to upload image during excel process: %s", e)
                        await msg.edit_text(f"❌ Failed to upload an image: {str(e)}")
                        return ConversationHandler.END
                # clear pending
                ctx.user_data['pending_images'] = []
                await msg.edit_text("⏳ Images uploaded! Processing spreadsheet data...")
                
            file = await ctx.bot.get_file(doc.file_id)
            byte_arr = await file.download_as_bytearray()
            
            rows = []
            if doc.file_name.endswith('.csv'):
                reader = csv.reader(io.StringIO(byte_arr.decode('utf-8')))
                rows = list(reader)
            else:
                wb = openpyxl.load_workbook(io.BytesIO(byte_arr), data_only=True)
                sheet = wb.active
                for r in sheet.iter_rows(values_only=True):
                    # Filter empty rows
                    if any(cell is not None for cell in r):
                        rows.append(r)
            
            if len(rows) < 2:
                await msg.edit_text("❌ File seems empty or has no data rows.")
                return ConversationHandler.END

            headers = [str(h).strip() if h else "" for h in rows[0]]
            data_rows = rows[1:]
            
            success = 0
            failed = 0
            errors = []
            
            for index, row in enumerate(data_rows, start=2):
                try:
                    name = self.extract_value(row, headers, ["product_name", "product name", "name"])
                    slug = self.extract_value(row, headers, ["slug"])
                        
                    if not name:
                        name = self.extract_value(row, headers, ["product_code", "product code", "code"]) # fallback
                        
                    if not name:
                        failed += 1
                        continue # Required
                        
                    if not slug:
                        slug = self.slugify(name)
                        
                    code = self.extract_value(row, headers, ["product_code", "product code", "code"])
                    price_str = self.extract_value(row, headers, ["discount", "price"])
                    orig_price_str = self.extract_value(row, headers, ["original"])
                    
                    def parse_price(pstr):
                        if not pstr: return 0
                        s = str(pstr).lower().replace('$', '').replace('n', '').replace(',', '').strip()
                        try: 
                            # Extract first number block
                            import re
                            nums = re.findall(r'\d+\.?\d*', s)
                            return float(nums[0]) if nums else 0
                        except: return 0
                    
                    price = parse_price(price_str)
                    orig_price = parse_price(orig_price_str)
                    
                    if price == 0 and orig_price > 0:
                        price = orig_price
                    
                    stock_str = self.extract_value(row, headers, ["stock"])
                    stock = 0
                    is_preorder = False
                    if stock_str:
                        slower = stock_str.lower()
                        if 'available' in slower:
                            stock = 99
                        elif 'pre' in slower:
                            is_preorder = True
                            stock = 0
                        else:
                            try:
                                import re
                                nums = re.findall(r'\d+', stock_str)
                                stock = int(nums[0]) if nums else 0
                            except: pass

                    capsize = self.split_commas_or_semicolons(self.extract_value(row, headers, ["capsize", "cap size"]))
                    bundles = self.split_commas_or_semicolons(self.extract_value(row, headers, ["bundles", "bundle"]))
                    colors = self.split_commas_or_semicolons(self.extract_value(row, headers, ["color", "colors"]))
                    parting = self.split_commas_or_semicolons(self.extract_value(row, headers, ["parting"]))
                    styling = self.split_commas_or_semicolons(self.extract_value(row, headers, ["styling", "style"]))
                    unavailable_lengths = self.split_commas_or_semicolons(self.extract_value(row, headers, ["unavailable lengths", "unavailable", "out of stock"]))
                    image_color_mapping_str = self.extract_value(row, headers, ["image color mapping", "color mapping"])
                    desc = self.extract_value(row, headers, ["description", "desc"])
                    
                    # Inches logic, split by spaces or commas
                    inches_str = self.extract_value(row, headers, ["inch", "length"])
                    inches = []
                    length_prices = []
                    
                    if inches_str:
                        if ',' in inches_str:
                            raw_inches = [x.strip() for x in inches_str.split(',') if x.strip()]
                        else:
                            raw_inches = [x.strip() for x in inches_str.split(' ') if x.strip()]
                            if len(raw_inches) > 1 and ',' not in inches_str:
                                raw_inches = [inches_str] 
                            else:
                                raw_inches = self.split_commas_or_semicolons(inches_str)
                                
                        for v in raw_inches:
                            if ':' in v or '$' in v:
                                parts = v.split(':') if ':' in v else v.split('$')
                                if len(parts) >= 2:
                                    import re
                                    length = parts[0].replace('$', '').strip()
                                    price_str = re.sub(r'[^\d.]', '', parts[1])
                                    if price_str:
                                        length_prices.append({"length": length, "price": float(price_str)})
                                        inches.append(length)
                                    else:
                                        inches.append(v)
                            else:
                                inches.append(v)

                    # description fallback if styling is present
                    final_desc = desc or ""
                    # We no longer need to append styling to desc since it has its own field,
                    # but keeping it for legacy or searchability if desired? 
                    # User specifically asked for styling field.
                    # if styling:
                    #     final_desc += f"\nStyling: {', '.join(styling)}"
                        
                    # Resolve category name → UUID
                    category_name = self.extract_value(row, headers, ["category", "cat"])
                    category_id = None
                    if category_name:
                        category_id = await api.resolve_category_id(category_name, self.token(ctx))
                        
                    # Always use only the images uploaded through telegram for this bulk insert
                    all_images = uploaded_image_urls
                        
                    payload = {
                        "name": name,
                        "product_name": name,
                        "slug": slug,
                        "product_code": code,
                        "price": price,
                        "original_price": orig_price if orig_price > price else None,
                        "stock": stock,
                        "is_preorder": is_preorder,
                        "cap_sizes": capsize if capsize else None,
                        "lengths": inches if inches else None,
                        "length_prices": length_prices if length_prices else None,
                        "bundles": bundles if bundles else None,
                        "colors": colors if colors else None,
                        "styling": styling if styling else None,
                        "unavailable_lengths": unavailable_lengths if unavailable_lengths else None,
                        "images": all_images if all_images else [],
                        "parting_options": parting if parting else None,
                        "description": final_desc.strip(),
                        "category_id": category_id,
                        "is_active": True,
                    }

                    # Handle image color mapping indices resolution
                    if image_color_mapping_str and uploaded_image_urls:
                        # expected format: "red":"1","blue":"2"
                        mapping_dict = {}
                        import re
                        pairs = re.findall(r'"([^"]+)":"([^"]+)"', image_color_mapping_str)
                        if not pairs:
                            # Try single pair without quotes if it failed e.g. Red:1
                            pairs = re.findall(r'([^:,]+):(\d+)', image_color_mapping_str)
                            
                        resolved_map = {}
                        for k, v in pairs:
                            try:
                                idx = int(v) - 1
                                if 0 <= idx < len(uploaded_image_urls):
                                    resolved_map[k.strip()] = uploaded_image_urls[idx]
                            except: pass
                        
                        if resolved_map:
                            payload["color_image_map"] = resolved_map
                    
                    await api.create_product(payload, self.token(ctx))
                    success += 1
                except Exception as e:
                    failed += 1
                    errors.append(f"Row {index} ({name}): {str(e)}")

            err_text = ""
            if errors:
                err_text = "\n\n*Errors:*\n" + "\n".join(errors[:5])
                if len(errors) > 5:
                    err_text += f"\n...and {len(errors)-5} more."
                    
            if success > 0 and failed == 0:
                header_msg = "✅ *Product Upload Complete*"
            elif success > 0 and failed > 0:
                header_msg = "⚠️ *Product Upload Partially Complete*"
            else:
                header_msg = "❌ *Product Upload Failed*"
                    
            await msg.edit_text(
                f"{header_msg}\n\n"
                f"• Successfully added product(s): {success}\n"
                f"• Failed: {failed}{err_text}",
                parse_mode="Markdown"
            )
        except Exception as e:
            log.error("Excel processing failed: %s", e)
            await msg.edit_text(f"❌ Failed to process file: {str(e)}")
            
        return ConversationHandler.END

    async def cancel(self, update: Update, ctx: ContextTypes.DEFAULT_TYPE) -> int:
        ctx.user_data['pending_images'] = []
        await update.message.reply_text("❌ Cancelled upload.")
        return ConversationHandler.END

    def build(self) -> ConversationHandler:
        return ConversationHandler(
            entry_points=[CommandHandler("bulkproduct", self.start)],
            states={
                WAIT_EXCEL: [
                    MessageHandler(filters.Document.ALL, self.handle_document),
                    MessageHandler(filters.PHOTO, self.handle_photo)
                ],
            },
            fallbacks=[CommandHandler("cancel", self.cancel)],
            allow_reentry=True,
        )
